# -*- coding = utf-8 -*-
# @Time : 2022/4/16 0:42
# @Author : 牧川
# @File : demo_xlwt.py
import xlwt
import openpyxl

# workbook = xlwt.Workbook(encoding='utf-8')  #创建wb对象
# worksheet = workbook.add_sheet('电影信息')   #创建工作表
# worksheet.write(0,1,'hello')               #向第1行第2列写入数据
# #写入九九乘法表
# for n in range(1,10):
#     for m in range(1,n+1):
#         worksheet.write(n,m, "%dx%d=%d"%(n,m,n*m))
#
# workbook.save('movie.xls')                #进行保存
from openpyxl import load_workbook
from openpyxl import Workbook

def open():
    #wb本身是一个迭代器，可以直接获取其中的工作表，也可以wb.sheetnames
    #wb = load_workbook('openyxl.xlsx')
    #激活工作表的几种方法，以下三种方法结果是同样的：
    sh1 = wb.active   #激活工作表，默认激活第一个
    #sh2 = wb['电影信息']    #直接按名称返回激活的工作表
    #sh3 = wb.get_sheet_by_name('电影信息')  #与上面类似，更推荐上面
    #获取单元格实例：
    v1 = sh1.cell(1,1)  #顺序是横坐标、纵坐标
    v1 = sh1['A1']  #直接通过索引获取单元格实例
    v2s = sh1['A1:B4']  #获取区域，返回单元格元组，可以用于迭代输出
    print(v1.value) #直接访问单元格实例的value属性
    #遍历一个表格：
    for row in sh1.rows:
        for cell in row:
            print(cell.value)
def creat():
    wb = Workbook()
    sheet = wb.active
    sheet2 = wb.create_sheet('sheet2')
    sheet['A1'] = '一个新的表格'
    #wb.save('openyxl.xlsx')
    #表格可以直接append列表
creat()
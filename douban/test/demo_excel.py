# -*- coding = utf-8 -*-
# @Time : 2022/5/2 13:22
# @Author : 牧川
# @File : demo_excel.py
from openpyxl import load_workbook
from openpyxl.styles import Font, colors, Alignment,PatternFill
from openpyxl import Workbook

#读写表格相关操作

def chai():
    wb = load_workbook('../豆瓣电影top250.xlsx')
    sh1 = wb.active
    sh2 = wb.create_sheet('低于9分')
    sh3 = wb.create_sheet('高于9分')
    dataLow = []
    dataUp = []
    for row in sh1.iter_rows(min_row=2):
        if row[4].value != '评分':
            if float(row[4].value) <9:
                dataLow.append(row)
            else:
                dataUp.append(row)

    for data in dataLow:
        dataList = []
        for v in data:
            dataList.append(v.value)
        print(dataList)
        sh2.append(dataList)
    for data in dataUp:
        dataList = []
        for v in data:
            dataList.append(v.value)
        sh3.append(dataList)

    wb.save('../豆瓣电影top250_2.xlsx')

#修改表格格式相关：
def changeFont():
    #获取表格对象
    wb = Workbook()
    sh1 = wb.active
    #实例化一个font对象：
    #常用的格式属性：字体、字号、加粗、颜色，颜色也可以直接用对应色号
    #myFont = Font(name='微软雅黑',size=30,bold=True,italic=True,color=colors.BLUE)
    myFont = Font(name='微软雅黑', size=30, bold=True, italic=True, color='FF6A6A')
    sh1['A1'].font = myFont

    #修改高度或者宽度：
    sh1.row_dimensions[6].height = 30   #行高
    sh1.column_dimensions['C'].width = 15   #列宽
    sh1['C6'] = 'hi,test my font'
    #修改对齐方式：
    sh1['C6'].alignment = Alignment(horizontal='right',vertical='top')

    #填充单元格背景：
    sh1['H6'].fill = PatternFill('solid','FF6D4C')

    wb.save('testFont.xlsx')
changeFont()